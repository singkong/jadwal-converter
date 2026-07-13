import re
import logging
from typing import Any
from openpyxl import load_workbook

import config
from models import ParseError

logger = logging.getLogger(__name__)


class ExcelReader:
    def __init__(self, filepath: str) -> None:
        self.filepath = filepath
        self.wb = load_workbook(filepath, data_only=True)
        self.ws = self.wb.active
        self.merged_map: dict[str, Any] = {}
        self._build_merged_map()

    def _build_merged_map(self) -> None:
        for merged_range in self.ws.merged_cells.ranges:
            min_col = merged_range.min_col
            min_row = merged_range.min_row
            max_col = merged_range.max_col
            max_row = merged_range.max_row
            top_left = self.ws.cell(row=min_row, column=min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    key = f"{row}:{col}"
                    self.merged_map[key] = top_left

    def cell_value(self, row: int, col: int) -> Any:
        key = f"{row}:{col}"
        if key in self.merged_map:
            return self.merged_map[key]
        return self.ws.cell(row=row, column=col).value

    @staticmethod
    def parse_jam(text: Any) -> tuple[str, str] | None:
        if text is None:
            return None
        text = str(text).replace(".", ":").strip()
        parts = re.split(r"\s*[-–—]\s*", text)
        if len(parts) != 2:
            return None
        mulai = parts[0].strip()
        selesai = parts[1].strip()
        if not mulai or not selesai:
            return None
        return (mulai, selesai)

    @staticmethod
    def parse_cell(text: Any, errors: list[ParseError] | None = None) -> tuple[int, str] | None:
        if text is None:
            return None
        raw = str(text).strip()
        if not raw:
            return None

        upper = raw.upper()
        for kw in config.IGNORE_KEYWORDS:
            if kw in upper:
                return None

        normalized = re.sub(r"[\n\r]+", " ", raw).strip()
        normalized = re.sub(r"\s+", " ", normalized)

        m = re.match(r"(\d+)\s*([A-Za-z])", normalized)
        if not m:
            return None

        teacher_code = int(m.group(1))
        subject_code = m.group(2).upper()

        tail = normalized[m.end():].strip()
        if tail and not re.match(r"^\s*$", tail):
            rest_upper = tail.upper()
            for kw in config.IGNORE_KEYWORDS:
                if kw in rest_upper:
                    return None
            m2 = re.match(r"\s*([A-Za-z])", tail)
            if m2:
                subject_code = m2.group(1).upper()

        return (teacher_code, subject_code)

    def get_info(self) -> dict[str, Any]:
        return {
            "sheet": self.ws.title,
            "rows": self.ws.max_row,
            "cols": self.ws.max_column,
        }

    def close(self) -> None:
        self.wb.close()