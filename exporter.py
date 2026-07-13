import csv
import logging
import sqlparse
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import config
from models import JadwalItem

logger = logging.getLogger(__name__)

BATCH_SIZE = 50


class Exporter:
    def __init__(self) -> None:
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)

    def export_sql(self, items: list[JadwalItem], filepath: str | Path) -> str:
        if not items:
            logger.warning("No items to export to SQL")
            return ""

        all_parts: list[str] = []
        for i in range(0, len(items), BATCH_SIZE):
            batch = items[i:i + BATCH_SIZE]
            sql = self._build_insert(batch)
            formatted = self._format_sql(sql)
            all_parts.append(formatted)

        final_sql = "\n\n".join(all_parts)

        filepath = Path(filepath)
        filepath.write_text(final_sql, encoding="utf-8")
        logger.info("SQL exported to %s (%d rows in %d batches)", filepath, len(items), len(all_parts))
        return final_sql

    def _build_insert(self, items: list[JadwalItem]) -> str:
        lines: list[str] = []
        lines.append("INSERT INTO `jadwals`")
        lines.append("(")
        cols = [
            "`tahun_ajaran_id`",
            "`guru_id`",
            "`kelas_id`",
            "`mapel_id`",
            "`hari`",
            "`jam_mulai`",
            "`jam_selesai`",
            "`created_at`",
            "`updated_at`",
        ]
        lines.append("  " + ", ".join(cols))
        lines.append(")")
        lines.append("VALUES")

        value_rows: list[str] = []
        for item in items:
            vals = (
                str(item.tahun_ajaran_id),
                str(item.guru_id),
                str(item.kelas_id),
                str(item.mapel_id),
                self._quote(item.hari),
                self._quote(item.jam_mulai),
                self._quote(item.jam_selesai),
                self._quote(item.created_at),
                self._quote(item.updated_at),
            )
            value_rows.append("  (" + ", ".join(vals) + ")")

        lines.append(",\n".join(value_rows))
        lines.append(";")
        return "\n".join(lines)

    def _format_sql(self, sql: str) -> str:
        try:
            return sqlparse.format(sql, reindent=True, keyword_case="upper")
        except Exception:
            return sql

    def export_excel(self, items: list[JadwalItem], filepath: str | Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Jadwal Import"

        headers = [
            "tahun_ajaran_id", "guru_id", "kelas_id", "mapel_id",
            "hari", "jam_mulai", "jam_selesai", "created_at", "updated_at",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.thin_border

        for row_idx, item in enumerate(items, 2):
            values = item.to_tuple()
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = self.thin_border
                if col_idx <= 4:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx >= 8:
                    cell.alignment = Alignment(horizontal="center")

        for col_idx in range(1, len(headers) + 1):
            max_len = len(headers[col_idx - 1])
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=len(items) + 1):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = max_len + 3

        filepath = Path(filepath)
        wb.save(str(filepath))
        logger.info("Excel exported to %s (%d rows)", filepath, len(items))

    def export_csv(self, items: list[JadwalItem], filepath: str | Path) -> None:
        headers = [
            "tahun_ajaran_id", "guru_id", "kelas_id", "mapel_id",
            "hari", "jam_mulai", "jam_selesai", "created_at", "updated_at",
        ]

        filepath = Path(filepath)
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for item in items:
                writer.writerow(item.to_tuple())

        logger.info("CSV exported to %s (%d rows)", filepath, len(items))

    @staticmethod
    def _quote(val: str) -> str:
        escaped = val.replace("'", "''")
        return f"'{escaped}'"